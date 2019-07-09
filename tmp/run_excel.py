import json

import requests
from openpyxl import load_workbook, Workbook


wb = load_workbook(filename='1.xlsx')
# print(wb.sheetnames)
sh = wb['添加加油卡']
# print(sh.cell(2, 2).value)
title_row = [sh.cell(row=1, column=i).value for i in range(1, sh.max_column+1)]
title = {v: i for i, v in enumerate(title_row)}
print(title)
data_rows = []
for row in sh.iter_rows(min_row=2):
    url = row[title['接口地址(名称)URL']].value
    method = row[title['方法']].value
    data = row[title['入参']].value

    res = requests.request(method, url, json=json.loads(data))
    row[title['测试结果']].value = res.text

wb.save('1.xlsx')


