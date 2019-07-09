import unittest
import json
import openpyxl
import ddt
import requests

wb = openpyxl.load_workbook("接口用例模板.xlsx")
sh = wb["添加加油卡"]

# print(sh[1])
# row_values = list(sh.rows)
title_data = [cell.value for cell in sh[1]]
# print(title_data)
# title_row = [sh.cell(1, i).value for i in range(1, sh.max_column)]
# print(title_row)
# for row in list(sh.rows)[1:]:
    # for cell in row:
        # print(cell.value, end=" ")
    # print()

# print([cell.value for row in sh.rows for cell in row])
case_data_list = []
for row in sh.iter_rows(min_row=2):
    row_data = [cell.value for cell in row]
    case_data = dict(zip(title_data, row_data))
    case_data_list.append(case_data)

# print(case_data_list)


def double_split(string, sep1, sep2):
    try:
        return {i.split(sep2)[0].strip(): i.split(sep2)[1].strip() for i in string.split(sep1)}
    except IndexError:
        raise ValueError("数据格式错误")
    

print(double_split('a : 1\nb: 2\n   c: 3  e: 2\n', "\n", ":"))

@ddt.ddt
class ApiTestCase(unittest.TestCase):
    @ddt.data(*case_data_list)
    def test_add_fuel_card(self, data):
        
        method = data.get("method", "get")
        url = data.get("url")
        params = data.get("params")


# unittest.main(verbosity=2)

# res = requests.request('POST', "http://www.baidu.com")

