import json
import openpyxl
import requests
from db import LongTengServer
from openpyxl.styles import Font, PatternFill, colors, Alignment

db = LongTengServer()

STATUS_STYLES = {
    'PASS': colors.GREEN,
    'ERROR': colors.YELLOW,
    'FAIL': colors.RED
}

class Excel(object):
    def __init__(self, file_path, sheet_name="执行用例"):
        self.wb = openpyxl.load_workbook(file_path)
        self.sh = self.wb.get_sheet_by_name(sheet_name)
        self.rows = list(self.sh.values)
        self.title_row = self.rows[0]
        self.status_index = self.title_row.index('STATUS') + 1
        self.error_msg_index = self.title_row.index('ERROR_MSG') + 1

    def get_sheet(self):
        rows = [dict(zip(self.title_row, row)) for row in self.rows[1:]]
        return rows

    def write_rows(self, cases):
        report_sh = self.wb.create_sheet(title='Report')

        report_sh.append(self.title_row)
        for case in cases:
            report_sh.append(list(case.values()))
        row = report_sh.row_dimensions[1]
        row.font = Font(underline="single", color=colors.RED)
        self.wb.save("report.xlsx")

    def write_results(self, cases):
        sh = self.wb.active
        start_row = 2
        ft = Font(color=colors.RED)
        for case in cases:
            status_cell = sh.cell(start_row, self.status_index)
            error_msg_cell = sh.cell(start_row, self.error_msg_index)
            status_cell.fill = PatternFill("solid", fgColor=STATUS_STYLES[case['STATUS']])
            status_cell.alignment = error_msg_cell.alignment = Alignment(horizontal="left", vertical="center")
            status_cell.value = case['STATUS']
            error_msg_cell.value = case['ERROR_MSG']

            start_row += 1
        self.wb.save("report.xlsx")


def json2dict(text):  # json字符串转字典
    text = text or '{}'  # 如果text为空, 则使用默认值 '{}'防止转字典报错
    try:
        return json.loads(text)
    except json.decoder.JSONDecodeError:
        print(f'{text}不是合法的json格式')


def run_setup(case):
    # 组装并执行setup
    setups = case['SETUP']
    if setups:  # 如果setups不为空
        for line in setups.split('\n'):  # 将setups按换行符分割,然后遍历
            print(f'  执行setup语句: {line.strip()}')
            eval(line.strip(), {}, {'db': db})


def send_reqeust(case):
    # 组装请求
    req = dict(
        method=case['METHOD'],
        url=case['URL'],
        params=json2dict(case['PARAMS']),
        headers=json2dict(case['HEADERS']),
        data=json2dict(case['DATA']),
        json=json2dict(case['JSON']))

    print(f'  发送请求: {req}')
    res = requests.request(**req).json()
    return res


def run_assert(case, res):
    ass = case['ASSERT']
    if ass:
        for line in ass.split('\n'):  # 将ass按换行符分割,然后遍历
            print(f'  执行断言语句: {line.strip()}')
            assert eval(line.strip(), {}, {'db': db, 'res': res})


def run_teardown(case):
    teardowns = case['TEARDOWN']
    if teardowns:  # 如果setups不为空
        for line in teardowns.split('\n'):  # 将setups按换行符分割,然后遍历
            print(f'  执行teardown语句: {line.strip()}')
            eval(line.strip(), {}, {'db': db})


def run_excel(file):
    excel = Excel(file)
    cases = excel.get_sheet()
    for case in cases:
        try:
            print(f"执行用例: 编号: {case['SN']} 模块: {case['MODULE']} 标题: {case['TITLE']} ")
            run_setup(case)
            res = send_reqeust(case)
            run_assert(case, res)
            run_teardown(case)   
        except AssertionError as ex:
            case['STATUS'] = 'FAIL'
            case['ERROR_MSG'] = str(ex)
        except Exception as ex:
            case['STATUS'] = 'ERROR'
            case['ERROR_MSG'] = str(ex)
        else:
            case['STATUS'] = 'PASS'
            case['ERROR_MSG'] = ''
    excel.write_results(cases)


if __name__ == '__main__':
    run_excel("data.xlsx")
